from pathlib import Path

import openpyxl

from rag.ingest import _strip_boilerplate, ingest_corpus, person_name, swap_name

SAMPLE = Path(__file__).resolve().parents[1] / "sample"


def _write_pdf(path: Path, pages: list[list[str]]) -> None:
    """Hand-build a minimal multi-page PDF: no reportlab, no pypdf writer needed.

    Each inner list is the text lines pypdf's `extract_text()` should read back for
    that page, one Tj per line at a fixed Helvetica position.
    """
    font_num = 3
    page_nums = [4 + 2 * i for i in range(len(pages))]
    content_nums = [5 + 2 * i for i in range(len(pages))]
    objects = {
        1: "<< /Type /Catalog /Pages 2 0 R >>",
        2: (
            f"<< /Type /Pages /Kids [{' '.join(f'{n} 0 R' for n in page_nums)}] "
            f"/Count {len(pages)} >>"
        ),
        font_num: "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    }
    for page_num, content_num, lines in zip(page_nums, content_nums, pages):
        objects[page_num] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 300 400] "
            f"/Resources << /Font << /F1 {font_num} 0 R >> >> /Contents {content_num} 0 R >>"
        )
        commands = ["BT /F1 12 Tf"]
        for row, line in enumerate(lines):
            escaped = line.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")
            commands.append(f"1 0 0 1 20 {380 - 16 * row} Tm ({escaped}) Tj")
        commands.append("ET")
        stream = "\n".join(commands)
        objects[content_num] = f"<< /Length {len(stream)} >>\nstream\n{stream}\nendstream"

    out = bytearray(b"%PDF-1.4\n")
    offsets = {}
    for num, body in sorted(objects.items()):
        offsets[num] = len(out)
        out += f"{num} 0 obj\n{body}\nendobj\n".encode("latin-1")
    xref_start = len(out)
    max_num = max(offsets)
    out += f"xref\n0 {max_num + 1}\n0000000000 65535 f \n".encode()
    for num in range(1, max_num + 1):
        out += f"{offsets.get(num, 0):010d} 00000 n \n".encode()
    out += f"trailer\n<< /Size {max_num + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF".encode()
    path.write_bytes(bytes(out))


def test_ingests_every_source_type():
    chunks = ingest_corpus(SAMPLE)
    types = {c.source_type for c in chunks}
    assert types == {"policy", "cv", "assignment", "credit", "aggregate"}


def test_person_name_survives_the_real_filename_shapes():
    assert person_name("Itenium - CV Alexander Ryckeboer") == "Alexander Ryckeboer"
    assert person_name("Itenium - CV Bernard Giorgino (FA)") == "Bernard Giorgino"
    assert person_name("Itenium - CV Bert Maes - Business Architect ") == "Bert Maes"
    assert person_name("Itenium - CV Bert Vermorgen - ENG") == "Bert Vermorgen"
    assert (
        person_name("Itenium - CV Bram De Plekker - .NET Angular Cloud Developer - updated")
        == "Bram De Plekker"
    )
    assert person_name("Itenium - CV Timo Versonnen- NL") == "Timo Versonnen"


def test_swap_name_turns_the_export_order_around():
    assert swap_name("Peeters, Dries") == "Dries Peeters"
    assert swap_name("De Plekker, Bram") == "Bram De Plekker"
    assert swap_name("Madonna") == "Madonna"


def test_assignment_chunks_drop_personal_data():
    chunks = [c for c in ingest_corpus(SAMPLE) if c.source_type == "assignment"]
    joined = "\n".join(c.text for c in chunks)
    # Assert on the field labels, not on bare values: "Antwerpen" is also a substring
    # of "UAntwerpen" in the College column, and "2000" is a zip code, a year and part
    # of "EUR 2000". A dropped column is one whose label never appears.
    for label in (
        "Birth Date", "Gender", "City", "State", "Zip Code", "Work Email", "LinkedIn URL",
    ):
        assert f"{label}:" not in joined
    # Two values that cannot collide with anything legitimate.
    for leaked in ("1990-04-11", "ana.meeus@example.be"):
        assert leaked not in joined
    assert "Klant: RetailCo" in joined


def test_credit_rows_are_one_chunk_each():
    chunks = [c for c in ingest_corpus(SAMPLE) if c.source_type == "credit"]
    assert len(chunks) == 28
    assert all(c.title for c in chunks)


def test_no_credit_chunk_contains_a_balance():
    """Question 5's whole point: the answer is in no retrievable chunk."""
    chunks = [c for c in ingest_corpus(SAMPLE) if c.source_type == "credit"]
    assert all("450" not in c.text for c in chunks)


def test_aggregate_chunk_states_the_summed_balance():
    chunks = [c for c in ingest_corpus(SAMPLE) if c.source_type == "aggregate"]
    dries = next(c for c in chunks if c.title == "Dries Peeters")
    assert "450" in dries.text


def test_aggregates_cover_every_person_in_the_ledger():
    chunks = ingest_corpus(SAMPLE)
    ledger_people = {c.title for c in chunks if c.source_type == "credit"}
    summary_people = {c.title for c in chunks if c.source_type == "aggregate"}
    assert ledger_people == summary_people


def test_cv_chunks_carry_the_person_name_as_title():
    titles = {c.title for c in ingest_corpus(SAMPLE) if c.source_type == "cv"}
    assert {"Dries Peeters", "Bram Claes"} <= titles


def test_policy_chunks_carry_a_heading_path():
    locations = [c.location for c in ingest_corpus(SAMPLE) if c.source_type == "policy"]
    assert any("Opleidingsbudget" in loc for loc in locations)


def test_chunk_ids_are_unique():
    chunks = ingest_corpus(SAMPLE)
    assert len({c.id for c in chunks}) == len(chunks)


# --- Defect 1: PDF chunks need a page citation ---


def test_pdf_chunks_get_one_location_per_page(tmp_path):
    (tmp_path / "pdfs").mkdir()
    _write_pdf(
        tmp_path / "pdfs" / "policy.pdf",
        [["Page one body text."], ["Page two body text."]],
    )
    locations = sorted(c.location for c in ingest_corpus(tmp_path))
    assert locations == ["policy > p. 1", "policy > p. 2"]


def test_pdf_chunk_text_lands_on_the_right_page(tmp_path):
    (tmp_path / "pdfs").mkdir()
    _write_pdf(
        tmp_path / "pdfs" / "policy.pdf",
        [["Page one body text."], ["Page two body text."]],
    )
    chunks = {c.location: c.text for c in ingest_corpus(tmp_path)}
    assert "Page one body text." in chunks["policy > p. 1"]
    assert "Page one body text." not in chunks["policy > p. 2"]


# --- Defect 2: repeated page furniture ---


def test_strip_boilerplate_removes_lines_on_at_least_half_the_pages():
    pages = [
        ["Header City BV", "Unique first page content"],
        ["Header City BV", "Unique second page content"],
        ["Header City BV", "Unique third page content"],
    ]
    stripped = _strip_boilerplate(pages)
    assert all("Header City BV" not in lines for lines in stripped)
    assert stripped[0] == ["Unique first page content"]


def test_strip_boilerplate_keeps_lines_under_half_the_pages():
    pages = [
        ["Only on page one", "Common line"],
        ["Common line"],
        ["Common line"],
    ]
    stripped = _strip_boilerplate(pages)
    assert stripped[0][0] == "Only on page one"


def test_strip_boilerplate_leaves_documents_under_three_pages_alone():
    pages = [["Repeated line"], ["Repeated line"]]
    assert _strip_boilerplate(pages) == pages


def test_strip_boilerplate_ignores_lines_shorter_than_three_characters():
    pages = [["12", "Body one"], ["12", "Body two"], ["12", "Body three"]]
    stripped = _strip_boilerplate(pages)
    assert all("12" in lines for lines in stripped)


def test_pdf_footer_is_stripped_from_ingested_chunks(tmp_path):
    (tmp_path / "pdfs").mkdir()
    footer = "ITENIUM BV BTW BE0000000000"
    _write_pdf(
        tmp_path / "pdfs" / "handbook.pdf",
        [
            ["Section one body.", footer],
            ["Section two body.", footer],
            ["Section three body.", footer],
        ],
    )
    chunks = ingest_corpus(tmp_path)
    assert all(footer not in c.text for c in chunks)
    assert any("Section one body." in c.text for c in chunks)


# --- Defect 3: spreadsheets were skipped ---


def test_xlsx_rows_become_chunks(tmp_path):
    (tmp_path / "pdfs").mkdir()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tarieven"
    sheet.append(["Type", "Bedrag"])
    sheet.append(["Auto", 0.42])
    workbook.save(tmp_path / "pdfs" / "rates.xlsx")

    chunks = [c for c in ingest_corpus(tmp_path) if c.source.endswith(".xlsx")]
    joined = "\n".join(c.text for c in chunks)
    assert "Type\tBedrag" in joined
    assert "Auto\t0.42" in joined


def test_xlsx_skips_empty_cells_and_rows(tmp_path):
    (tmp_path / "pdfs").mkdir()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Type", None, "Bedrag"])
    sheet.append([None, None, None])
    sheet.append(["Auto", None, 0.42])
    workbook.save(tmp_path / "pdfs" / "rates.xlsx")

    chunks = [c for c in ingest_corpus(tmp_path) if c.source.endswith(".xlsx")]
    joined = "\n".join(c.text for c in chunks)
    assert "Type\tBedrag" in joined
    assert "Auto\t0.42" in joined
    assert "\t\t" not in joined


def test_xlsx_sheets_keep_their_own_name_and_rows(tmp_path):
    (tmp_path / "pdfs").mkdir()
    workbook = openpyxl.Workbook()
    jan = workbook.active
    jan.title = "Jan"
    jan.append(["Day", "Kms"])
    jan.append(["1", "12"])
    feb = workbook.create_sheet("Feb")
    feb.append(["Day", "Kms"])
    feb.append(["1", "34"])
    workbook.save(tmp_path / "pdfs" / "kms.xlsx")

    chunks = [c for c in ingest_corpus(tmp_path) if c.source.endswith(".xlsx")]
    by_location = {c.location: c.text for c in chunks}
    assert "12" in by_location["kms > Jan"]
    assert "34" in by_location["kms > Feb"]
    assert "12" not in by_location["kms > Feb"]
    assert "34" not in by_location["kms > Jan"]


def test_blank_xlsx_produces_no_chunks(tmp_path):
    (tmp_path / "pdfs").mkdir()
    workbook = openpyxl.Workbook()
    workbook.save(tmp_path / "pdfs" / "template.xlsx")

    chunks = [c for c in ingest_corpus(tmp_path) if c.source.endswith(".xlsx")]
    assert chunks == []
