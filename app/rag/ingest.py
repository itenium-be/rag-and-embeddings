"""Parse before chunking. Each source type gets the treatment its shape deserves."""

from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from dataclasses import asdict
from pathlib import Path

from rag.chunking import chunk_id, split_text
from rag.models import Chunk

# Both spellings are supported: `sample/` and the real `data/raw/` use the same names,
# and `policies/` stays accepted so an English-named drop still works.
PROSE_DIRS = [("policy", ["pdfs", "policies"]), ("cv", ["cvs"]), ("project", ["projects"])]

# Date of birth and home address are on the leave-out list in the spec, and no question
# needs them. Dropping at ingest means they never reach the vector store at all.
ASSIGNMENT_DROP = {
    "Birth Date", "Gender", "City", "State", "Zip Code", "Country",
    "Work Email", "LinkedIn URL", "First Name", "Middle Name",
}


def read_document(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        from pypdf import PdfReader

        return "\n\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
    if path.suffix.lower() == ".docx":
        import docx

        return "\n\n".join(p.text for p in docx.Document(str(path)).paragraphs)
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for sheet in workbook.worksheets:
            rows = [
                "\t".join(str(cell) for cell in row if cell not in (None, ""))
                for row in sheet.iter_rows(values_only=True)
            ]
            rows = [row for row in rows if row]
            # "##", not "#": `_sections` keeps one doc_title variable, overwritten by
            # every "#" line and applied to ALL sections at the end — one "#" per
            # sheet would mislabel every sheet but the last with the wrong name.
            if rows:
                sheets.append(f"## {sheet.title}\n" + "\n".join(rows))
        if not sheets:
            return ""
        return f"# {path.stem}\n\n" + "\n\n".join(sheets)
    return path.read_text(encoding="utf-8")


def person_name(filename: str) -> str:
    """Pull the consultant's name out of a CV filename.

    The real exports are named "Itenium - CV Bram De Plekker - .NET Angular Cloud
    Developer - updated.pdf". Everything after the first " - " is role, language or
    revision noise, and a trailing "(FA)" marks a variant of the same person.
    """
    stem = re.sub(r"^itenium\s*-\s*cv\s*", "", filename, flags=re.IGNORECASE)
    stem = stem.split(" - ")[0]
    stem = re.sub(r"\([^)]*\)", "", stem)
    return stem.strip() or filename


def swap_name(name: str) -> str:
    """"Peeters, Dries" -> "Dries Peeters", so records read like the CVs do."""
    if "," not in name:
        return name.strip()
    last, first = name.split(",", 1)
    return f"{first.strip()} {last.strip()}".strip()


def _sections(text: str) -> list[tuple[str, str]]:
    """Split markdown on headings, returning (heading path, body) pairs."""
    doc_title = ""
    sections: list[tuple[str, list[str]]] = []
    current = ""
    body: list[str] = []
    for line in text.splitlines():
        match = re.match(r"^(#{1,6})\s+(.*)$", line)
        if match:
            if body:
                sections.append((current, body))
                body = []
            heading = match.group(2).strip()
            if len(match.group(1)) == 1:
                doc_title = heading
                current = ""
            else:
                current = heading
        else:
            body.append(line)
    if body:
        sections.append((current, body))
    return [
        (f"{doc_title} > {head}" if head else doc_title, "\n".join(lines).strip())
        for head, lines in sections
        if "\n".join(lines).strip()
    ]


def _document_title(text: str, fallback: str) -> str:
    match = re.search(r"^#\s+(.*)$", text, re.MULTILINE)
    if not match:
        return fallback
    # CVs are titled "Name — Role"; the name alone is what a citation should say.
    return match.group(1).split("—")[0].strip()


def _strip_boilerplate(pages: list[list[str]]) -> list[list[str]]:
    """Drop lines that repeat across at least half a document's pages: running headers
    and footers, not content — real content varies page to page. Purely structural
    (counts repetition, never looks at the text itself) so it works on any PDF.
    """
    if len(pages) < 3:
        return pages
    counts: dict[str, int] = defaultdict(int)
    for lines in pages:
        for stripped in {line.strip() for line in lines if len(line.strip()) >= 3}:
            counts[stripped] += 1
    boilerplate = {line for line, count in counts.items() if 2 * count >= len(pages)}
    return [[line for line in lines if line.strip() not in boilerplate] for lines in pages]


def _pdf_pages(path: Path, title: str) -> list[tuple[str, str]]:
    from pypdf import PdfReader

    pages = [(page.extract_text() or "").splitlines() for page in PdfReader(path).pages]
    pages = _strip_boilerplate(pages)
    return [
        (f"{title} > p. {n}", body)
        for n, lines in enumerate(pages, start=1)
        if (body := "\n".join(lines).strip())
    ]


def _location_pairs(path: Path, text: str, title: str) -> list[tuple[str, str]]:
    """(location, body) pairs for one document — how it's carved up depends on format."""
    if path.suffix.lower() == ".pdf":
        return _pdf_pages(path, title)
    return _sections(text)


def _prose_chunks(path: Path, root: Path, source_type: str) -> list[Chunk]:
    text = read_document(path)
    if not text.strip():
        return []
    source = str(path.relative_to(root))
    fallback = person_name(path.stem) if source_type == "cv" else path.stem
    title = _document_title(text, fallback)
    chunks: list[Chunk] = []
    for heading_path, body in _location_pairs(path, text, title):
        for part in split_text(body):
            chunks.append(
                Chunk(
                    id=chunk_id(source, len(chunks), part),
                    text=part,
                    source=source,
                    source_type=source_type,
                    title=title,
                    location=heading_path or title,
                )
            )
    return chunks


def _rows(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _assignment_chunks(path: Path, root: Path) -> list[Chunk]:
    """One row per assignment, rendered as fields.

    This is the wrong way to index structured data, and that is the point: the records
    come out as near-identical blobs sitting on top of each other in vector space.
    """
    source = str(path.relative_to(root))
    chunks: list[Chunk] = []
    for ordinal, row in enumerate(_rows(path)):
        name = swap_name(row.get("Last name, First name", ""))
        fields = [f"Naam: {name}"] + [
            f"{key}: {value}"
            for key, value in row.items()
            if key and key not in ASSIGNMENT_DROP and key != "Last name, First name" and value
        ]
        text = "\n".join(fields)
        chunks.append(
            Chunk(
                id=chunk_id(source, ordinal, text),
                text=text,
                source=source,
                source_type="assignment",
                title=name,
                location="BambooHR opdracht",
            )
        )
    return chunks


def _credit_chunks(path: Path, root: Path) -> list[Chunk]:
    source = str(path.relative_to(root))
    chunks: list[Chunk] = []
    for ordinal, row in enumerate(_rows(path)):
        name = swap_name(row.get("Last name, First name", ""))
        text = (
            f"Naam: {name}\n"
            f"Datum: {row.get('Effective Date', '')}\n"
            f"Event: {row.get('Event', '')}\n"
            f"Type: {row.get('Event Type', '')}\n"
            f"Credits: {row.get('Credits', '')}"
        )
        chunks.append(
            Chunk(
                id=chunk_id(source, ordinal, text),
                text=text,
                source=source,
                source_type="credit",
                title=name,
                location=f"Creditsboeking {row.get('Effective Date', '')}",
            )
        )
    return chunks


def _credit_aggregate_chunks(path: Path, root: Path) -> list[Chunk]:
    """The answer, computed at ingest time.

    Nothing here is a retrieval technique. Summing a ledger is arithmetic over records,
    which is exactly what vector search cannot do — so it happens before the vectors
    exist. Hidden until wizard step 6.
    """
    source = str(path.relative_to(root))
    earned: dict[str, float] = defaultdict(float)
    spent: dict[str, float] = defaultdict(float)
    events: dict[str, int] = defaultdict(int)
    latest: dict[str, str] = defaultdict(str)

    for row in _rows(path):
        name = swap_name(row.get("Last name, First name", ""))
        try:
            credits = float(row.get("Credits") or 0)
        except ValueError:
            continue
        (earned if credits >= 0 else spent)[name] += credits
        events[name] += 1
        latest[name] = max(latest[name], row.get("Effective Date") or "")

    chunks: list[Chunk] = []
    for ordinal, name in enumerate(sorted(events)):
        balance = earned[name] + spent[name]
        text = (
            f"Creditsaldo voor {name}.\n"
            f"Huidig saldo: {balance:g} credits.\n"
            f"Verdiend: {earned[name]:g}. Ingeruild: {abs(spent[name]):g}.\n"
            f"Aantal boekingen: {events[name]}. Laatste boeking: {latest[name] or 'onbekend'}."
        )
        chunks.append(
            Chunk(
                id=chunk_id(f"{source}#aggregate", ordinal, text),
                text=text,
                source=source,
                source_type="aggregate",
                title=name,
                location="Berekend creditsaldo",
            )
        )
    return chunks


def _bamboo_dir(root: Path) -> Path | None:
    directory = root / "bamboo"
    return directory if directory.is_dir() else None


def ingest_corpus(root: Path) -> list[Chunk]:
    root = Path(root)
    chunks: list[Chunk] = []

    for source_type, names in PROSE_DIRS:
        for name in names:
            directory = root / name
            if not directory.is_dir():
                continue
            for path in sorted(directory.iterdir()):
                if path.suffix.lower() in {".md", ".txt", ".pdf", ".docx", ".xlsx"}:
                    chunks.extend(_prose_chunks(path, root, source_type))

    bamboo = _bamboo_dir(root)
    if bamboo:
        for path in sorted(bamboo.glob("*.csv")):
            if "credit" in path.stem.lower():
                chunks.extend(_credit_chunks(path, root))
                chunks.extend(_credit_aggregate_chunks(path, root))
            else:
                chunks.extend(_assignment_chunks(path, root))

    return chunks


def write_chunks(chunks: list[Chunk], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for chunk in chunks:
            handle.write(json.dumps(asdict(chunk), ensure_ascii=False) + "\n")


def read_chunks(path: Path) -> list[Chunk]:
    with Path(path).open(encoding="utf-8") as handle:
        return [Chunk(**json.loads(line)) for line in handle if line.strip()]
